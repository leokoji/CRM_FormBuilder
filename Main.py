from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse
import sqlite3
from openpyxl import Workbook
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, numbers

html = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Registro</title>

        <style>
        input, textarea, select {
            width: 300px;
            padding: 5px;
            margin-bottom: 10px;
        }

        body{
            background: #f2f4f6;
            font-family: Arial, Helvetica, sans-serif;
            margin: 20px;
            color: #2c3e50;
            align-items: flex;
        }

        header{
            background: #2c3e50;
            color: #ffff;
            padding: 14px 24px;
            display:flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 24px;
            border-radius: 10px;
        }

        header h1{
            font-size: 18px;
            margin: 0;
        }

        .container{
            max-width: 600px
            margin: 0 auto;
            padding: 0 20px;
        }

        .card{
            background: #fff;
            border-radius: 6px;
            padding: 24px;
            box-shadow: 0 1px 3px rgba(0,0,0,1);
        }

        button.acao{
            background: #2c3e50;
            color: #fff;
            border: none;
            padding: 10px 18px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 14px;
        }

        button.acao:hover{
            background: #2980b9;
        }

        label{
            display: block;
            margin-bottom: 4px;
            font-weight: bold;
        }

        input, textarea, select{
            width: 100%;
            padding: 8px;
            margin-bottom: 16px;
            border: 1px solid #ccd1d9;
            border-radius: 4px;
            font-size: 14px;
            box-sizing: border-box;
        }
    </style>
</head>
<body>

    <header>
        <h1>CRM FORMBUILDER</h1>
    </header>

    <div class="container">
        <div class="card">
            <form action="/enviar" method="post">
            <div>
                <label for="empresa">Empresa</label>
                <input name="empresa" type="text" id="empresa" required>
            </div>

            <div>
                <label for="vidas">Vidas</label>
                <input name="vidas" type="number" id="vidas" required>
            </div>

            <div>
                <label for="status">Status</label>
                <select name="status" id="status" required>
                    <option value="Agendado">Agendado</option>
                    <option value="Realizado">Realizado</option>
            </select>
            </div>

            <div>
                <label for="obs">Observação</label>
                <textarea name="observacao" id="obs" rows="3"></textarea>
            </div>

            <div>
                <label for="contato">Contato</label>
                <input name="contato" type="text" id="contato" required>
            </div>

            <div>
                <label for="telefone">Telefone</label>
                <input name="telefone" type="tel" id="telefone" required>
            </div>

            <div>
                <label for="email">Email</label>
                <input name="email" type="email" id="email" required>
            </div>

            <div>
                <label for="consultora">Consultora</label>
                <input name="consultora" type="text" id="consultora" required>
            </div>

            <div>
                <label for="analista">Analista</label>
                <input name="analista" type="text" id="analista" required>
            </div>

            <div>
                <label for="congenere">Congenere</label>
                <input name="congenere" type="text" id="congenere" required>
            </div>

            <div>
                <label for="hotphone">HotPhone</label>
                <select name="hotphone" id="hotphone" required>
                    <option value="Sim">Sim</option>
                    <option value="Nao">Não</option>
            </select>
            </div>

            <div>
                <label for="agendamento">Data De Agendamento</label>
                <input name="agendamento" type="date" id="agendamento" required>
            </div>

            <div>
                <label for="visita">Data De Visita</label>
                <input name="visita" type="date" id="visita" required>
            </div>

            <div>
                <label for="horario">Horario da visita</label>
                <input name="horario" type="time" id="horario" required>
            </div>


            <div style="margin-top: 20px; display: flex; gap: 10px;">
                <button type="submit" class="acao">Enviar</button>
                <a href="/exportar"><button type="button" class="acao">Exportar Excel</button></a>
            </div>
        </form>
    </div>
</body>
</html>
"""

conn = sqlite3.connect("Leads.db")

conn.execute("""CREATE TABLE IF NOT EXISTS agendamentos (
             id INTEGER PRIMARY KEY AUTOINCREMENT,
             agendamento TEXT,
             visita TEXT,
             horario TEXT,
             empresa TEXT,
             contato TEXT,
             email TEXT,
             vidas INTEGER,
             congenere TEXT,
             hotphone TEXT,
             status TEXT,
             observacao TEXT,
             analista TEXT, 
             consultora TEXT,
             telefone TEXT,
             criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

conn.execute("""CREATE TABLE IF NOT EXISTS usuario (
             id INTEGER PRIMARY KEY AUTOINCREMENT,
             nome TEXT,
             senha TEXT
             criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

app = FastAPI()


def formatar_planilha(ws):

    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import Alignment
    from datetime import datetime

    tab = Table(displayName="Agendamentos", ref=ws.dimensions)

    style = TableStyleInfo(
        name="TableStyleMedium7",
        showRowStripes=True,
    )

    tab.tableStyleInfo = style
    ws.add_table(tab)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and len(cell.value) >= 10:
                try:
                    dt = datetime.strptime(cell.value[:10], "%Y-%m-%d")
                    cell.value = dt
                    cell.number_format = "DD/MM/YYYY"
                except (ValueError, TypeError):
                    pass
@app.get("/Login")
def login(nome: str = Form(...), senha: str = Form(...)):
    conn = sqlite3.connect("Leads.db")
    return HTMLResponse(content=html)

@app.get("/Form")
def main():
    return HTMLResponse(content=html)

@app.post("/enviar")
def receber_dados(
    agendamento: str = Form(...),
    visita: str = Form(...),
    horario: str = Form(...),
    empresa: str = Form(...),
    contato: str = Form(...),
    email: str = Form(...),
    vidas: int = Form(...),
    congenere: str = Form(...),
    hotphone: str = Form(...),
    status: str = Form(...),
    observacao: str = Form(...),
    analista: str = Form(...),
    consultora: str = Form(...),
    telefone: str = Form(...)):

    conn = sqlite3.connect("Leads.db")

    conn.execute("""INSERT INTO agendamentos (
                 agendamento,
                 visita,
                 horario,
                 empresa,
                 contato,
                 email,
                 vidas,
                 congenere,
                 hotphone,
                 status,
                 observacao,
                 analista,
                 consultora,
                 telefone)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (agendamento, visita, horario, empresa, contato, email, vidas, congenere, hotphone, status, observacao, analista, consultora, telefone))
    
    conn.commit()
    conn.close()
    return HTMLResponse(content="<h2>Dados enviados com sucesso!</h2><a href='/'>Novo registro</a>")

@app.get("/exportar")
def exportar():
    conn = sqlite3.connect("Leads.db")
    rows = conn.execute("SELECT * FROM agendamentos").fetchall()
    conn.close()

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Agendamentos"

    itens = ["ID", "Agendamento", "Visita", "Horario", "Empresa", "Contato", "Email", "Vidas", "Congenere", "HotPhone", "Status", "Observacao", "Analista", "Consultora", "Telefone", "Criado Em"]
    ws.append(itens)

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    formatar_planilha(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
       output,
       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
       headers={"content-disposition": "attatchment; filename=agendamentos.xlsx"})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
